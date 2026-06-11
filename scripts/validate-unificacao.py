#!/usr/bin/env python3
"""Valida se data/planilha/unificacao.json bate com U_DinamicaColada da planilha."""

from __future__ import annotations

import json
import sys
from datetime import timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path(
    r"c:\Users\vinicius.morais\Documents\Unattend Bots\Tommy\Relatórios\Principal_11_06.xlsm"
)
JSON_PATH = ROOT / "data" / "planilha" / "unificacao.json"

COLUMNS = [
    ("tangerino.horasNormais", "Soma de T_Horas Normais", 1),
    ("tangerino.horasExtras", "Soma de T_Horas Extras", 2),
    ("tangerino.sobreaviso", "Soma de T_Sobreaviso", 3),
    ("tangerino.total", "Soma de Tangerino Total", 4),
    ("orange.horasNormais", "Soma de F_Horas Normais", 5),
    ("orange.horasExtras", "Soma de F_Horas Extras", 6),
    ("orange.sobreaviso", "Soma de F_Sobreaviso", 7),
    ("orange.total", "Soma de FcTeam_Total", 8),
    ("graficos.tSobreavisoTerco", "T_Sobreaviso 1/3", 9),
    ("graficos.fSobreavisoTerco", "F_Sobreaviso_1/3", 10),
    ("graficos.tHorasExtras", "T_Horas Extras (grafico)", 11),
    ("graficos.fHorasExtras", "F_Horas Extras (grafico)", 12),
]


def td_to_hours(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, timedelta):
        return round(value.total_seconds() / 3600, 2)
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    return 0.0


def read_excel_unificacao(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    ws = None
    for name in wb.sheetnames:
        if "dinamicacol" in name.lower().replace(" ", ""):
            ws = wb[name]
            break
    if ws is None:
        raise RuntimeError("Aba U_DinamicaColada não encontrada")

    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        nome = str(row[0]).strip()
        item: dict = {"analista": nome, "tangerino": {}, "orange": {}, "graficos": {}}
        for key, _, idx in COLUMNS:
            parts = key.split(".")
            item[parts[0]][parts[1]] = td_to_hours(row[idx] if idx < len(row) else None)
        rows.append(item)
    wb.close()
    return rows


def get_nested(d: dict, path: str) -> float:
    cur = d
    for p in path.split("."):
        cur = cur[p]
    return float(cur)


def main() -> int:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    if not source.exists():
        print(f"ERRO: planilha não encontrada: {source}")
        return 1
    if not JSON_PATH.exists():
        print(f"ERRO: JSON não encontrado: {JSON_PATH}")
        return 1

    excel = read_excel_unificacao(source)
    exported = json.loads(JSON_PATH.read_text(encoding="utf-8"))

    excel_map = {r["analista"]: r for r in excel}
    json_map = {r["analista"]: r for r in exported}

    errors = []
    print(f"Planilha: {source.name}")
    print(f"Analistas Excel: {len(excel_map)} | JSON: {len(json_map)}")
    print()

    missing_in_json = set(excel_map) - set(json_map)
    missing_in_excel = set(json_map) - set(excel_map)
    if missing_in_json:
        errors.append(f"Faltando no JSON: {sorted(missing_in_json)}")
    if missing_in_excel:
        errors.append(f"Sobrando no JSON: {sorted(missing_in_excel)}")

    for nome in sorted(excel_map):
        if nome not in json_map:
            continue
        ex = excel_map[nome]
        js = json_map[nome]
        for key, label, _ in COLUMNS:
            ev = get_nested(ex, key)
            jv = get_nested(js, key)
            if abs(ev - jv) > 0.01:
                errors.append(f"{nome} | {label}: Excel={ev} JSON={jv} (diff={jv-ev:+.2f})")

    if errors:
        print("DIVERGÊNCIAS ENCONTRADAS:")
        for e in errors:
            print(f"  - {e}")
        print(f"\nTotal: {len(errors)} problema(s)")
        return 1

    print("OK — todos os valores de U_DinamicaColada batem com unificacao.json")
    for nome in sorted(excel_map):
        ex = excel_map[nome]
        print(
            f"  {nome}: T={ex['tangerino']['total']}h | O={ex['orange']['total']}h"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
