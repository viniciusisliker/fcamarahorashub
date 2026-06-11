#!/usr/bin/env python3
"""Exporta dados da planilha Principal (Tommy) para JSON consumido pelo hub."""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "data" / "planilha"
DEFAULT_SOURCE = Path(
    r"c:\Users\vinicius.morais\Documents\Unattend Bots\Tommy\Relatórios\Principal_11_06.xlsm"
)


def sheet_by_keyword(wb: openpyxl.Workbook, *keywords: str) -> openpyxl.worksheet.worksheet.Worksheet:
    for name in wb.sheetnames:
        lower = name.lower()
        if all(k.lower() in lower for k in keywords):
            return wb[name]
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return wb[name]
    raise KeyError(f"Nenhuma aba encontrada para: {keywords}")


def td_to_hours(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, timedelta):
        return round(value.total_seconds() / 3600, 2)
    if isinstance(value, (int, float)):
        return round(float(value) * 24, 2) if float(value) < 5 else round(float(value), 2)
    return 0.0


def parse_br_decimal(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return round(float(text), 2)
    except ValueError:
        return 0.0


def parse_br_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def slug_id(*parts: str) -> str:
    base = "-".join(p for p in parts if p)
    base = re.sub(r"[^a-zA-Z0-9_-]+", "-", base).strip("-").lower()
    return base[:120] or "item"


def export_analistas(wb: openpyxl.Workbook) -> list[dict]:
    ws = sheet_by_keyword(wb, "analistas")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    result = []
    for row in rows:
        if not row or not row[0]:
            continue
        result.append(
            {
                "nome": str(row[0]).strip(),
                "email": str(row[1]).strip() if row[1] else "",
                "cargo": str(row[2]).strip() if row[2] else "",
                "responsavel": str(row[3]).strip() if row[3] else "",
                "status": str(row[4]).strip() if row[4] else "Ativo",
            }
        )
    return result


def export_unificacao(wb: openpyxl.Workbook) -> list[dict]:
    ws = sheet_by_keyword(wb, "dinamicacolada")
    headers = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col_map = {h: i for i, h in enumerate(headers)}

    def col(*names: str) -> int | None:
        for name in names:
            for h, i in col_map.items():
                if name.lower() in h.lower():
                    return i
        return None

    idx_nome = col("analista") or 0
    idx_t_total = col("tangerino total", "soma de tangerino total")
    idx_f_total = col("fcteam_total", "soma de fcteam_total")
    idx_t_sobre = col("t_sobreaviso") if col("t_sobreaviso") != col("soma de t_sobreaviso") else col("soma de t_sobreaviso")
    idx_f_sobre = col("f_sobreaviso") if col("f_sobreaviso") else col("soma de f_sobreaviso")
    idx_t_extra = col("t_horas extras", "soma de t_horas extras")
    idx_f_extra = col("f_horas extras", "soma de f_horas extras")
    idx_t_norm = col("t_horas normais", "soma de t_horas normais")
    idx_f_norm = col("f_horas normais", "soma de f_horas normais")

  # Extra columns used directly by charts (J,K,L,M in sheet)
    extra_cols = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    chart_t_sobre = 9 if len(extra_cols) > 9 else None
    chart_f_sobre = 10 if len(extra_cols) > 10 else None
    chart_t_extra = 11 if len(extra_cols) > 11 else None
    chart_f_extra = 12 if len(extra_cols) > 12 else None

    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx_nome]:
            continue
        nome = str(row[idx_nome]).strip()
        result.append(
            {
                "analista": nome,
                "tangerino": {
                    "horasNormais": td_to_hours(row[idx_t_norm]) if idx_t_norm is not None else 0,
                    "horasExtras": td_to_hours(row[chart_t_extra]) if chart_t_extra is not None else td_to_hours(row[idx_t_extra]) if idx_t_extra is not None else 0,
                    "sobreaviso": td_to_hours(row[chart_t_sobre]) if chart_t_sobre is not None else td_to_hours(row[idx_t_sobre]) if idx_t_sobre is not None else 0,
                    "total": td_to_hours(row[idx_t_total]) if idx_t_total is not None else 0,
                },
                "orange": {
                    "horasNormais": td_to_hours(row[idx_f_norm]) if idx_f_norm is not None else 0,
                    "horasExtras": td_to_hours(row[chart_f_extra]) if chart_f_extra is not None else td_to_hours(row[idx_f_extra]) if idx_f_extra is not None else 0,
                    "sobreaviso": td_to_hours(row[chart_f_sobre]) if chart_f_sobre is not None else td_to_hours(row[idx_f_sobre]) if idx_f_sobre is not None else 0,
                    "total": td_to_hours(row[idx_f_total]) if idx_f_total is not None else 0,
                },
            }
        )
    return result


def export_apontamentos_por_dia(wb: openpyxl.Workbook) -> dict:
    ws = sheet_by_keyword(wb, "colad")
    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    dias = [int(v) for v in header[1:] if isinstance(v, (int, float))]
    analistas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        valores = []
        for v in row[1 : 1 + len(dias)]:
            valores.append(td_to_hours(v) if v else 0)
        analistas.append({"analista": str(row[0]).strip(), "dias": valores})
    return {"dias": dias, "analistas": analistas}


def export_timesheet(wb: openpyxl.Workbook, analistas: list[dict]) -> list[dict]:
    ws = wb["F_Timesheet"]
    email_equipe = {a["email"].lower(): a.get("responsavel", "Hyper") for a in analistas if a.get("email")}
    nome_equipe = {a["nome"].lower(): a.get("responsavel", "Hyper") for a in analistas}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    result = []
    for i, row in enumerate(rows):
        if not row or not row[1]:
            continue
        data_iso = parse_br_date(row[1])
        if not data_iso:
            continue
        nome = str(row[7]).strip() if row[7] else "Desconhecido"
        email = str(row[8]).strip().lower() if row[8] else ""
        worker = str(row[0]).strip() if row[0] else email or nome
        horas = parse_br_decimal(row[12]) if row[12] is not None else td_to_hours(row[11])
        if horas <= 0:
            continue
        equipe = email_equipe.get(email) or nome_equipe.get(nome.lower(), "Hyper")
        tarefa = str(row[5]).strip() if row[5] else ""
        anotacoes = str(row[18]).strip() if len(row) > 18 and row[18] and row[18] != "--" else ""
        descricao = anotacoes or tarefa or "Apontamento Orange/FCTeam"
        result.append(
            {
                "id": slug_id(worker, data_iso, str(row[9] or i), str(i)),
                "colaboradorId": worker,
                "colaboradorNome": nome,
                "equipe": equipe,
                "data": data_iso,
                "projeto": str(row[4]).strip() if row[4] else "—",
                "cliente": str(row[2]).strip() if row[2] and row[2] != "--" else None,
                "horas": horas,
                "descricao": descricao[:500],
                "status": "aprovado",
                "fonte": "orange",
            }
        )
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.source.exists():
        print(f"Arquivo não encontrado: {args.source}", file=sys.stderr)
        return 1

    print(f"Lendo {args.source}...")
    wb = openpyxl.load_workbook(args.source, read_only=True, data_only=True, keep_vba=True)

    analistas = export_analistas(wb)
    unificacao = export_unificacao(wb)
    por_dia = export_apontamentos_por_dia(wb)
    apontamentos = export_timesheet(wb, analistas)
    wb.close()

    args.output.mkdir(parents=True, exist_ok=True)

    meta = {
        "exportadoEm": datetime.now().isoformat(),
        "arquivoOrigem": args.source.name,
        "totalApontamentos": len(apontamentos),
        "totalAnalistas": len(analistas),
        "periodo": {
            "inicio": min((a["data"] for a in apontamentos), default=None),
            "fim": max((a["data"] for a in apontamentos), default=None),
        },
    }

    files = {
        "meta.json": meta,
        "analistas.json": analistas,
        "unificacao.json": unificacao,
        "apontamentos-por-dia.json": por_dia,
        "apontamentos.json": apontamentos,
    }

    for name, payload in files.items():
        path = args.output / name
        with path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        print(f"  {name}: {path.stat().st_size // 1024} KB")

    print("Exportação concluída.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
